from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from src.screener.engine import run_preset


OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "screener_output.xlsx"


PRESETS = [
    "Quality Compounder",
    "Value Pick",
    "Growth Accelerator",
    "Dividend Champion",
    "Debt-Free Blue Chip",
    "Turnaround Watch",
]


# ---------------------------------------------------------
# Columns required in the final screener report
# ---------------------------------------------------------

KPI_COLUMNS = [
    "company_id",
    "year",
    "broad_sector",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "free_cash_flow_cr",
    "cfo_quality_score",
    "fcf_conversion_rate",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "debt_to_equity",
    "interest_coverage",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "market_cap_crore",
    "composite_quality_score",
]


# ---------------------------------------------------------
# Preset threshold definitions
# Used only for Excel colour coding
# ---------------------------------------------------------

PRESET_RULES = {
    "Quality Compounder": {
        "return_on_equity_pct": ("min", 15),
        "debt_to_equity": ("max", 1.0),
        "free_cash_flow_cr": ("min", 0),
        "revenue_cagr_5yr": ("min", 10),
    },

    "Value Pick": {
        "pe_ratio": ("max", 20),
        "pb_ratio": ("max", 3.0),
        "debt_to_equity": ("max", 2.0),
        "dividend_yield_pct": ("min", 1),
    },

    "Growth Accelerator": {
        "pat_cagr_5yr": ("min", 20),
        "revenue_cagr_5yr": ("min", 15),
        "debt_to_equity": ("max", 2.0),
    },

    "Dividend Champion": {
        "dividend_yield_pct": ("min", 2),
        "dividend_payout_ratio_pct": ("max", 80),
        "free_cash_flow_cr": ("min", 0),
    },

    "Debt-Free Blue Chip": {
        "debt_to_equity": ("max", 0),
        "return_on_equity_pct": ("min", 12),
    },

    "Turnaround Watch": {
        "revenue_cagr_3yr": ("min", 10),
        "free_cash_flow_cr": ("min", 0),
    },
}


# ---------------------------------------------------------
# Helper: Find the correct column name
# ---------------------------------------------------------

def find_column(df, column_name):
    if column_name in df.columns:
        return column_name

    alternatives = {
        "pe_ratio": ["pe", "p_e", "pe_ratio"],
        "pb_ratio": ["pb", "p_b", "pb_ratio"],
        "market_cap_crore": [
            "market_cap",
            "market_cap_cr",
            "market_cap_crore",
        ],
    }

    for alternative in alternatives.get(column_name, []):
        if alternative in df.columns:
            return alternative

    return None


# ---------------------------------------------------------
# Check whether a cell passes a preset rule
# ---------------------------------------------------------

def passes_rule(value, operator, threshold):

    if pd.isna(value):
        return False

    if operator == "min":
        return value >= threshold

    if operator == "max":
        return value <= threshold

    return False


# ---------------------------------------------------------
# Generate Excel report
# ---------------------------------------------------------

def generate_screener_output():

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    all_results = {}

    for preset in PRESETS:

        print(f"\nGenerating: {preset}")

        try:
            df = run_preset(preset)
        except Exception as exc:
            print(f"ERROR: {preset}: {exc}")
            continue

        if df is None:
            df = pd.DataFrame()

        if not df.empty:

            # Sort by composite score
            if "composite_quality_score" in df.columns:
                df = df.sort_values(
                    by="composite_quality_score",
                    ascending=False,
                )

            # Keep available KPI columns
            available_columns = [
                col
                for col in KPI_COLUMNS
                if col in df.columns
            ]

            # Add any important columns missing from KPI list
            for col in df.columns:
                if col not in available_columns:
                    if col not in [
                        "id",
                        "company_id",
                        "year",
                    ]:
                        available_columns.append(col)

            df = df[available_columns]

        all_results[preset] = df

        print(
            f"{preset}: "
            f"{len(df)} companies"
        )


    # -----------------------------------------------------
    # Write all 6 sheets
    # -----------------------------------------------------

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        for preset, df in all_results.items():

            # Excel sheet names max 31 characters
            sheet_name = preset[:31]

            if df.empty:
                df = pd.DataFrame(
                    {
                        "Message": [
                            "No companies matched this preset."
                        ]
                    }
                )

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )


    # -----------------------------------------------------
    # Apply formatting
    # -----------------------------------------------------

    workbook = load_workbook(OUTPUT_FILE)

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for preset in PRESETS:

        sheet_name = preset[:31]

        if sheet_name not in workbook.sheetnames:
            continue

        ws = workbook[sheet_name]

        # Header formatting
        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # Freeze header
        ws.freeze_panes = "A2"

        # Find columns
        column_map = {}

        for cell in ws[1]:
            column_map[
                cell.value
            ] = cell.column

        # -------------------------------------------------
        # Colour-code preset threshold cells
        # -------------------------------------------------

        rules = PRESET_RULES.get(
            preset,
            {},
        )

        for metric, rule in rules.items():

            actual_column = find_column(
                pd.DataFrame(
                    columns=column_map.keys()
                ),
                metric,
            )

            if actual_column is None:
                continue

            if metric not in column_map:
                continue

            column_number = column_map[metric]

            operator, threshold = rule

            for row in range(
                2,
                ws.max_row + 1,
            ):

                cell = ws.cell(
                    row=row,
                    column=column_number,
                )

                try:
                    value = float(
                        cell.value
                    )
                except (
                    TypeError,
                    ValueError,
                ):
                    continue

                if passes_rule(
                    value,
                    operator,
                    threshold,
                ):
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

        # -------------------------------------------------
        # Auto-width columns
        # -------------------------------------------------

        for column_cells in ws.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            ws.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                30,
            )


    workbook.save(OUTPUT_FILE)

    print("\n" + "=" * 60)
    print(
        "SCREENER EXCEL REPORT GENERATED"
    )
    print("=" * 60)

    print(
        f"File: {OUTPUT_FILE}"
    )

    print(
        f"Sheets: {len(workbook.sheetnames)}"
    )

    for sheet in workbook.sheetnames:
        print(
            f"  - {sheet}"
        )

    print("=" * 60)


if __name__ == "__main__":
    generate_screener_output()